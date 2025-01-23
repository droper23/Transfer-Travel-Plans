import customtkinter
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import Calendar
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import re
import os
import time

def initialize_driver(): # Initialize Chrome web driver
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver

def get_home(folder_name): # Get home directory
    home_dir = os.path.expanduser("~") # Define folder path
    folder_path = os.path.join(home_dir, folder_name)
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Folder '{folder_name}' created at {folder_path}")
    else:
        print(f"Folder '{folder_name}' exists at {folder_path}")
    
    return folder_path

main_folder = get_home("GoogleMapsTrains")  # Creates 'GoogleMapsTrains' folder
output_folder = get_home(os.path.join("GoogleMapsTrains", "Screenshots"))  # Creates the screenshot folder within the GoogleMapsTrains folder
save_directory = main_folder

print(f"Main folder: {main_folder}")
print(f"Screenshot folder: {output_folder}")
print(f"Save directory: {save_directory}") # Should be the same as the Main Folder one
os.makedirs(save_directory, exist_ok=True)

start_city = "Keleti pályaudvar" # Define intermediate city that trains are going into/out of. In this case it's just Keleti Station in Budapest.

missionaries = [
{"Name": "Person A", "Email": "example_a@example.com"},
{"Name": "Person B", "Email": "example_b@example.com"},
{"Name": "Person C", "Email": "example_c@example.com"},
{"Name": "Person D", "Email": "example_d@example.com"},
]

missionary_names = [missionary["Name"] for missionary in missionaries]
cities = ["Sopron", "Szombathely", "Pápa", "Veszprém", "Győr", "Kaposvár", "Pécs", "Székesfehérvár", "Dunaújváros", "Kecskemét", "Szeged", "Szolnok", "Békéscsaba", "Eger", "Miskolc", "Nyíregyháza", "Debrecen", "Pest", "Buda", "Óbuda"]

customtkinter.set_appearance_mode("system")
customtkinter.set_default_color_theme("blue")

def debug_log(message): # Prints debug log
    print(f"[DEBUG] {message}")

app = customtkinter.CTk()
app.title("Google Maps Trains")
app.geometry("485x600")

departure_date_var = tk.StringVar()
departure_time_var = tk.StringVar()

def pick_date():
    def set_date():
        departure_date_var.set(calendar.get_date())
        print(f"[DEBUG] Selected Date: {departure_date_var.get()}")
        date_window.destroy()

    date_window = customtkinter.CTkToplevel(app)
    date_window.title("Select Date")

    main_window_x = app.winfo_x()
    main_window_y = app.winfo_y()
    main_window_width = app.winfo_width()
    date_window.geometry(f"+{main_window_x + main_window_width + 10}+{main_window_y}") # Sets the position of the date window directly to the right of the main window

    calendar = Calendar(date_window, date_pattern="yyyy-mm-dd", background="#1f6aa5", foreground="white", selectbackground="#1f6aa5", selectforeground="white", headersbackground="#1f6aa5", headersforeground="white", normalbackground="lightblue", normalforeground="black", weekendbackground="lightblue", weekendforeground="black")
    calendar.pack(padx=10, pady=10)
    select_button = customtkinter.CTkButton(date_window, text="Select", command=set_date).pack(pady=10)

customtkinter.CTkLabel(app, text="Transfer Date:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
date_button = customtkinter.CTkButton(app, text="Pick Date", command=pick_date)
date_button.grid(row=0, column=1, padx=10, pady=5)
date_label = customtkinter.CTkLabel(app, textvariable=departure_date_var)
date_label.grid(row=0, column=2, padx=10, pady=5, sticky="w")

def update_time_label(value): # Creates slider to choose the transfer time. Default time is 11 am
    hours = int(value) // 2 + 6
    minutes = int(value) % 2 * 30
    period = "AM" if hours < 12 else "PM"
    hours = hours if hours <= 12 else hours - 12
    departure_time_var.set(f"{hours:02}:{minutes:02} {period}")
default_time_value = 5 * 2

customtkinter.CTkLabel(app, text="Transfer Time:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
time_slider = customtkinter.CTkSlider(app, from_=8, to=24, command=update_time_label)
time_slider.set(default_time_value)
time_slider.grid(row=1, column=1, padx=10, pady=5)

update_time_label(default_time_value)
time_label = customtkinter.CTkLabel(app, textvariable=departure_time_var)
time_label.grid(row=1, column=2, padx=10, pady=5, sticky="w")

missionary_frame = customtkinter.CTkFrame(app)
missionary_frame.grid(row=2, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")
missionary_rows = []

def add_missionary_row():
    row = {}
    name_var = tk.StringVar()
    from_var = tk.StringVar()
    to_var = tk.StringVar()

    row_index = len(missionary_rows) + 2

    name_dropdown = customtkinter.CTkComboBox(missionary_frame, variable=name_var, values=missionary_names, state="readonly")
    name_dropdown.grid(row=row_index, column=0, padx=5, pady=2)

    from_dropdown = customtkinter.CTkComboBox(missionary_frame, variable=from_var, values=cities, state="readonly")
    from_dropdown.grid(row=row_index, column=1, padx=5, pady=2)

    to_dropdown = customtkinter.CTkComboBox(missionary_frame, variable=to_var, values=cities, state="readonly")
    to_dropdown.grid(row=row_index, column=2, padx=5, pady=2)

    row["name"] = name_var
    row["from"] = from_var
    row["to"] = to_var

    def get_selected_cities():  # Returns the selected cities as str
        from_city = from_var.get() 
        to_city = to_var.get()
        return f"From: {from_city}, To: {to_city}"
    missionary_rows.append(row)
    print(get_selected_cities())

customtkinter.CTkLabel(missionary_frame, text="Missionary Name").grid(row=1, column=0, padx=5, pady=5)
customtkinter.CTkLabel(missionary_frame, text="From City").grid(row=1, column=1, padx=5, pady=5)
customtkinter.CTkLabel(missionary_frame, text="To City").grid(row=1, column=2, padx=5, pady=5)

add_row_button = customtkinter.CTkButton(app, text="Add Missionary", command=add_missionary_row)
add_row_button.grid(row=3, column=0, columnspan=2, padx=10, pady=10)

departure_time_obj = datetime.strptime(departure_time_var.get(), "%H:%M %p")
out_departure_time_obj = departure_time_obj + timedelta(minutes=30) # Sets departure time from intermediate city to 30 mins. after the set transfer time
out_departure_time = out_departure_time_obj.strftime("%H:%M")

def select_departure_date(driver, target_date):
    try:
        print(f"[DEBUG] Raw target_date: {target_date}")
        match = re.match(r"(\d{4})-(\d{2})-(\d{2})", target_date)
        
        if match:
            target_year = match.group(1)
            target_month = match.group(2)
            target_day = match.group(3) 
            month_map = {
                "01": "jan.",
                "02": "feb.",
                "03": "mar.",
                "04": "apr.",
                "05": "may",
                "06": "jun.",
                "07": "jul.",
                "08": "aug.",
                "09": "sep.",
                "10": "oct.",
                "11": "nov.",
                "12": "dec."
            }
            target_month_abbreviation = month_map.get(target_month, f"Unknown ({target_month})")

            print(f"[DEBUG] Target Day: {target_day}, Target Month: {target_month_abbreviation}, Target Year: {target_year}")
        else:
            print("[DEBUG] Error: Unable to match date pattern.")
            return

        date_picker_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//span[@aria-live='polite']"))
        )
        date_picker_button.click()

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//table[@role='presentation']"))
        )
        aria_label = f"{target_day} {target_month_abbreviation}"
        
        date_cell = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, f"//td[@role='gridcell' and @aria-label='{aria_label}']"))
        )
        date_cell.click()

    except Exception as e:
        print(f"[DEBUG] Error selecting date: {e}")

def handle_cookies(driver):
    try:
        cookies_button = WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "button[aria-label^='Az összes elfogadása']"))
        )
        
        if cookies_button.is_displayed():
            cookies_button.click()
            debug_log("Cookies accepted")
        else:
            debug_log("Cookies buttons not found/already clicked")
    except Exception as e:
        debug_log(f"Cookies button not found/already clicked: {e}")

def search_transportation(driver, start, destination, departure_time, departure_date, arrival=False): # Search Google Maps for transportation data
    base_url = "https://www.google.com/maps"
    driver.get(base_url)

    handle_cookies(driver)

    try:
        search_box = WebDriverWait(driver, 15).until( # Input destination
            EC.presence_of_element_located((By.ID, "searchboxinput"))
        )
        time.sleep(3)  # Allow autoinput to finish
        search_box.clear()
        search_box.send_keys(destination)
        search_box.send_keys(Keys.RETURN)
        debug_log(f"Destination '{destination}' entered in search bar.")

        route_button = WebDriverWait(driver, 15).until( # Wait for "Útvonalterv" button then click it
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button[aria-label^='Útvonalterv']"))
        )
        route_button.click()
        debug_log("Clicked 'Útvonalterv' button.")

        start_box = WebDriverWait(driver, 15).until( # Wait for the searchbox to input start location
            EC.presence_of_element_located((By.CSS_SELECTOR, "input.tactile-searchbox-input"))
        )
        time.sleep(2)  # Let autoinput finish
        start_box.clear()
        start_box.send_keys(start)
        start_box.send_keys(Keys.RETURN)
        debug_log(f"Start location '{start}' entered.")

        transport_button = WebDriverWait(driver, 15).until( # Click public transportation button
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-tooltip='Tömegközlekedés']"))
        )
        transport_button.click()
        debug_log("Clicked 'Tömegközlekedés' button.")

        departure_dropdown = WebDriverWait(driver, 15).until( # Open thing to select "Indulás" or "Érkezés"
            EC.element_to_be_clickable((By.CSS_SELECTOR, "div[aria-label='Indulási beállítások']"))
        )
        departure_dropdown.click()
        debug_log("Dropdown for 'Indulási beállítások' clicked.")
        time.sleep(1)  # Wait for dropdown to fully expand

        option_text = "Érkezés" if arrival else "Indulás" # Determine whether its érkezés or indulás
        debug_log(f"Determined dropdown selection: {option_text}")

        try: # Wait for the érkezés/indulás option to be visible and select it
            option_element = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, f"//div[contains(@class, 'goog-menuitem-content') and text()='{option_text}']"))
            )
            option_element.click()
            debug_log(f"'{option_text}' successfully selected.")
        except Exception as e:
            debug_log(f"Failed to select '{option_text}': {e}")
            driver.save_screenshot("dropdown_debug.png")
            debug_log("Saved screenshot for dropdown failure.")

        
        time_input = WebDriverWait(driver, 15).until( # Input departure/arrival time
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='transit-time']"))
        )
        time_input.clear()
        time_input.send_keys(departure_time)
        time_input.send_keys(Keys.RETURN)
        debug_log(f"Time '{departure_time}' entered and submitted.")

        select_departure_date(driver, departure_date) # References other function to input departure date

        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1.VuCHmb.fontHeadlineSmall#section-directions-trip-title-0"))
            )
            debug_log("Search results loaded successfully.")
        except TimeoutException:
            debug_log("Search results did not load within the specified time.")

    except Exception as e:
        debug_log(f"Error during search: {e}")

def extract_transport_info(driver, from_city, to_city): # Pull all the travel data from the search
    transport_data = {
        "From": from_city,
        "To": to_city,
        "Travel Time": "N/A",
        "Departure Time": "N/A",
        "Screenshot": "N/A"
    }

    try:
        try: # Extract travel time
            travel_time_element = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.Fk3sm.fontHeadlineSmall"))
            )
            transport_data["Travel Time"] = travel_time_element.text
        except TimeoutException:
            debug_log("Travel Time element not found; using default 'N/A'.")

        try: # Extract departure time
            departure_time_element = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.tUEI8e.fontBodyMedium span span"))
            )
            transport_data["Departure Time"] = departure_time_element.text
        except TimeoutException:
            debug_log("Departure Time element not found; using default 'N/A'.")

        sidebar_region = (65, 150, 402, 800)  # Sets the size of the sidebar to be sc-d. Adjust these based on your screen.
        screenshot_filename = capture_sidebar_by_pixels(driver, f"{from_city}_to_{to_city}", sidebar_region)
        if screenshot_filename:
            transport_data["Screenshot"] = screenshot_filename

    except Exception as e:
        debug_log(f"Error extracting transport info: {e}")

    debug_log(f"Extracted data for {to_city}: {transport_data}")
    return transport_data

def capture_sidebar_by_pixels(driver, filename_prefix, region): # Get a screenshot of the sidebar based on pixels
    try:
        details_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button[aria-labelledby='section-directions-trip-details-msg-0']"))
        )
        details_button.click()
        debug_log("Clicked 'Részletek' button to expand details.")

        time.sleep(2)  # Let the sidebar expand. Edit to be longer if sc's are weird

        full_screenshot_filename = os.path.join(output_folder, f"{filename_prefix}_full_page.png") # Take a full-page screenshot
        driver.save_screenshot(full_screenshot_filename)
        debug_log(f"Full-page screenshot saved as '{full_screenshot_filename}'.")

        with PILImage.open(full_screenshot_filename) as img: # Crop the full page sc to just the side bar
            left, top, width, height = region
            cropped_img = img.crop((left, top, left + width, top + height))
            cropped_screenshot_filename = os.path.join(output_folder, f"{filename_prefix}_sidebar.png")
            cropped_img.save(cropped_screenshot_filename)
            debug_log(f"Cropped sidebar screenshot saved as '{cropped_screenshot_filename}'.")

            return cropped_screenshot_filename

    except Exception as e:
        debug_log(f"Error capturing sidebar screenshot: {e}")
        return None

def save_to_excel(data, filename="transportation_schedule.xlsx", header_color="CADAD7", odd_row_color="F5F1EE", even_row_color="FFFFFF"): # Save the data to an Excel file
    if not data:
        print("No data to save.")
        return

    file_path = os.path.join(main_folder, filename) 

    wb = Workbook()
    ws = wb.active
    ws.title = "Transportation Options"

    headers = ["Direction", "From", "To", "Travel Time", "Departure Time", "Screenshot"]
    ws.append(headers)

    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid") # Style the header
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True, color="405955")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill

    for i, entry in enumerate(data, start=2): # Add all the info to the excel sheet
        from_city = entry.get("From", "N/A")
        to_city = entry.get("To", "N/A")

        if from_city.lower().startswith("keleti") or "budapest" in from_city.lower(): # Determine direction
            direction = f"To: {to_city}"
        else:
            direction = f"From: {from_city}"

        row_data = [
            direction,                  # Direction
            from_city,                  # From
            to_city,                    # To
            entry.get("Travel Time", "N/A"),  # Travel Time
            entry.get("Departure Time", "N/A"),  # Departure Time
            ""                          # Screenshot
        ]
        ws.append(row_data)

        
        row_fill = PatternFill( # Alternate row colors
            start_color=odd_row_color if i % 2 == 0 else even_row_color,
            end_color=odd_row_color if i % 2 == 0 else even_row_color,
            fill_type="solid"
        )
        for cell in ws[i]:
            cell.fill = row_fill

        
        screenshot_filename = entry.get("Screenshot", None) 
        if screenshot_filename and os.path.exists(screenshot_filename): # Embed the screenshot image 
            img = ExcelImage(screenshot_filename)
            img.height = 450  # Adjust height for the image. Edit nyugodtan for better excel formating
            img.width = 200   # Adjust width for the image. Edit nyugodtan for better excel formating
            ws.add_image(img, f"F{i}")  # Place the image in the Screenshot column (column F)

    for col in ws.columns: # Auto-adjust column widths
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    try:
        wb.save(file_path)
        print(f"Data successfully saved to {file_path}")
    except Exception as e:
        print(f"Failed to save the file: {e}")

def start_search():
    global selected_cities # Get the selected cities from the dropdowns
    selected_cities = [] 

    for row in missionary_rows: # Loop through each missionary and get the selected cities
        from_city = row["from"].get()  
        to_city = row["to"].get() 
        
        if from_city not in selected_cities: 
            selected_cities.append(from_city)
        if to_city not in selected_cities:
            selected_cities.append(to_city)

    destination_cities = selected_cities # Create destination cities from selected cities without duplicates

    print("Selected Cities for Search:")
    for city in selected_cities:
        print(city)
    print(f"Destination Cities: {destination_cities}")

    departure_date = departure_date_var.get() # Get departure date and time
    if not departure_date:
        print("[DEBUG] Departure Date is empty.")
        messagebox.showerror("Error", "Please select a departure date.")
        return
    departure_time = departure_time_var.get()
    driver = initialize_driver()

    if not departure_date or not departure_time:
        messagebox.showerror("Error", "Please select both departure date and time.")
        return

    departure_time_obj = datetime.strptime(departure_time_var.get(), "%I:%M %p") # Convert time to 24-hour time
    
    out_departure_time_obj = departure_time_obj + timedelta(minutes=30) # Calculate departure time from arrival time. Edit the '30' to change the difference between arrival and departure time.
    out_departure_time = out_departure_time_obj.strftime("%H:%M")

    debug_log(f"Selected departure time: {departure_time}, Out departure time: {out_departure_time}")

    driver = initialize_driver() # Initialize the driver and perform the search
    try:
        all_data = []

        for city in destination_cities: # Collect data for trips from Keleti pályaudvar to cities
            search_transportation(driver, "Keleti pályaudvar", city, out_departure_time, departure_date)
            city_data = extract_transport_info(driver, "Keleti pályaudvar", city)
            all_data.append(city_data)

            search_transportation(driver, city, "Budapest-Keleti, Budapest, Kerepesi út 2-4, 1087", departure_time, departure_date, arrival=True) # Collect data for trips from cities to Keleti pályaudvar
            reverse_city_data = extract_transport_info(driver, city, "Budapest-Keleti, Budapest, Kerepesi út 2-4, 1087")
            all_data.append(reverse_city_data)

        debug_log(f"Collected data: {all_data}")
        save_to_excel(all_data)

    finally:
        driver.quit()

start_button = customtkinter.CTkButton(app, text="Start Search", command=start_search) # Button to start the search
start_button.grid(row=5, column=1, padx=10, pady=10)

app.mainloop() # Start the Tkinter main loop

if __name__ == "__main__":
    driver = initialize_driver()
    destination_cities = selected_cities

    try:
        all_data = []

        for city in destination_cities: # Collect data for trips from Keleti pályaudvar to cities
            search_transportation(driver, "Keleti pályaudvar", city, out_departure_time, departure_date_var)
            city_data = extract_transport_info(driver, "Keleti pályaudvar", city)
            all_data.append(city_data)

        for city in destination_cities: # Collect data for trips from cities to Keleti pályaudvar
            search_transportation(driver, city, "Budapest-Keleti, Budapest, Kerepesi út 2-4, 1087", departure_time_var, departure_date_var, arrival=True)
            reverse_city_data = extract_transport_info(driver, city, "Budapest-Keleti, Budapest, Kerepesi út 2-4, 1087")
            all_data.append(reverse_city_data)

        debug_log(f"Collected data: {all_data}")
        print(f"Current working directory: {os.getcwd()}")
        
        save_to_excel(all_data)

    finally:
        driver.quit()
